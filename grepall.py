import os, PyPDF2, re, argparse

from twkr import *

from openpyxl import load_workbook

from termcolor import colored

parser = argparse.ArgumentParser()

parser.add_argument("-i", "--path", help="Path", default=".")

parser.add_argument("-d", "--depth", help="Depth", default="max")

parser.add_argument("-p", "--patterns", help="Pattern(s) to search ex: 'yes,no'", default="")

parser.add_argument("-r", "--get_row", help="get the row of the pattern's location", type=bool, default=False)

parser.add_argument("-e", "--file_dir_to_exclude", help="all the files or folders not to be searched ex: 'dir/a.txt,b.txt'", default="")

parser.add_argument("-E", "--ptrn_excl", help="all the pattern that are contained in the folders or files not to be searched ex: 'ir,.txt,.pdf'", default="")

args = parser.parse_args()

##### CONF VAR ######

libre_unoconv = [".ppt", ".docx", ".odt", ".pptx"]

pre_ptrn_excl = [".jpg", ".jpeg", ".gif", ".png", ".mp4", ".mp3", ".avi", ".mkv", ".tar", ".zip", ".dvi", ".gz", ".log"] # etc

####################

path = args.path

depth = args.depth # int for particular

list_to_search = args.patterns.split(",")

if list_to_search[0] == "" and len(list_to_search) == 1:

    list_to_search = []

get_row = args.get_row # may slow down the programm

file_dir_to_exclude = args.file_dir_to_exclude.split(",")

if file_dir_to_exclude[0] == "" and len(file_dir_to_exclude) == 1:

    file_dir_to_exclude = []

ptrn_excl = args.ptrn_excl.split(",")

if ptrn_excl[0] == "" and len(ptrn_excl) == 1:

    ptrn_excl = []

[ ptrn_excl.append(ext) for ext in pre_ptrn_excl ]

def grep_xlsx(list_to_search, file):

    wb = load_workbook('teste.xlsx')

    l_sheet = wb.sheetnames

    l_all = []

    l_limit = []

    l_limitc = []

    l_limitr = []

    for i in range(0, len(l_sheet)):

        ws = wb[l_sheet[i]]

        min_row = ws.min_row

        min_column = ws.min_column

        max_row = ws.max_row

        max_column = ws.max_column

        for ic in range (min_column, max_column + 1):

            for ir in range(min_row, max_row + 1):

                current = ws.cell(ir, ic).value

                if current != None:

                    l_all.append(str(current))

                    l_limitc.append(ic)

                    l_limitr.append(ir)

        l_limit.append(len(l_all))

    stop = 0

    for i in range (0, len(list_to_search)):

        for i2 in range (0, len(l_all)):

            if list_to_search[i] == l_all[i2]:

                for i3 in range(0, len(l_limit)):

                    if i2 < l_limit[i3] and stop == 0:

                        nb_s = i3

                        stop = 1

                stop = 0

                print("Match found in sheet", nb_s + 1, "at column", l_limitc[i2], "at row", l_limitr[i2], "with", colored(list_to_search[i]), "red")

def grep_raw(inpt_f, list_to_search):

    word_re =  "(" + "|".join(list_to_search) + ")"

    with open(inpt_f, "r") as f:

        cnt = f.read()

    if get_row == True: 

        cnt = cnt.split("\n")

        for row in cnt:

            pre_l = list(set(re.findall(word_re, row)))

            if pre_l:

                [ print("The pattern", colored("{}".format(ptrn), "red"), "has been found at row {}".format(row)) for ptrn in pre_l ]

    else:

        pre_l = list(set(re.findall(word_re, cnt)))

        if pre_l:

            [ print("The pattern", colored("{}".format(ptrn), "red"), "has been found") for ptrn in pre_l ]

def grep_pdf(inpt_f, list_to_search):

    pdf = PyPDF2.PdfReader(inpt_f)

    pgs = pdf.pages

    word_re = "(" + "|".join(list_to_search) + ")"

    for i in range(len(pgs)):

        pre_l = list(set(re.findall(word_re, pgs[i].extract_text())))

        if pre_l:

            [ print("The pattern", colored("{}".format(ptrn), "red"), "has been found at page {}".format(i)) for ptrn in pre_l ]

all_f = file_rec(path=path, tracker_l=[os.listdir(path)], depth=depth, excl=file_dir_to_exclude, sub_excl=ptrn_excl, frst_path=path)

for file in all_f:

    if "." in file:

        nf = None

        print("")

        print("For file", colored("{}".format(file), "cyan"))

        ext = file[file.index("."):]

        f_ = path + "/" + file

        if ext in libre_unoconv:

            print("")

            os.system("unoconv -f pdf '{}'".format(f_))

            f_ = list(f_)

            f_[-len(libre_unoconv[libre_unoconv.index(ext)]):] = ".pdf"

            f_ = "".join(f_)

            nf = f_

            print(f_)

            ext = ".pdf"

        if ext == ".xlsx":

            grep_xlsx(list_to_search=list_to_search, file=f_)

        elif ext == ".pdf":

            grep_pdf(inpt_f=f_, list_to_search=list_to_search)

        else:

            grep_raw(inpt_f=f_, list_to_search=list_to_search)

        if nf != None:

            os.remove(nf)

print("")

print("done")

